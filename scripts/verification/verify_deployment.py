#!/usr/bin/env python3
"""Aave V4 deployment verification script.

Reads a deployment report JSON and a config input JSON, connects to an RPC
endpoint, and verifies that all on-chain state matches the expected configuration.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import time
from dataclasses import dataclass, field
from decimal import Decimal
from functools import lru_cache
from pathlib import Path
from typing import Any, Optional
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from web3 import Web3

ARTIFACTS_DIR = Path(__file__).resolve().parent.parent.parent / "out"


def load_abi(sol_file: str, contract_name: str) -> list:
    path = ARTIFACTS_DIR / sol_file / f"{contract_name}.json"
    if not path.exists():
        sys.exit(f"Artifact not found: {path}\nRun `forge build` first.")
    with open(path) as f:
        return json.load(f)["abi"]


@dataclass(frozen=True)
class ArtifactInfo:
    sol_file: str
    contract_name: str
    impl_sol_file: Optional[str] = None
    impl_contract_name: Optional[str] = None


ARTIFACT_MAP: dict[str, ArtifactInfo] = {
    "AccessManager": ArtifactInfo("AccessManagerEnumerable.sol", "AccessManagerEnumerable"),
    "HubConfigurator": ArtifactInfo("HubConfigurator.sol", "HubConfigurator"),
    "SpokeConfigurator": ArtifactInfo("SpokeConfigurator.sol", "SpokeConfigurator"),
    "TreasurySpoke": ArtifactInfo(
        "TransparentUpgradeableProxy.sol", "TransparentUpgradeableProxy",
        impl_sol_file="TreasurySpokeInstance.sol",
        impl_contract_name="TreasurySpokeInstance",
    ),
    "Hub": ArtifactInfo(
        "TransparentUpgradeableProxy.sol", "TransparentUpgradeableProxy",
        impl_sol_file="HubInstance.sol",
        impl_contract_name="HubInstance",
    ),
    "InterestRateStrategy": ArtifactInfo(
        "AssetInterestRateStrategy.sol", "AssetInterestRateStrategy",
    ),
    "Spoke": ArtifactInfo(
        "TransparentUpgradeableProxy.sol", "TransparentUpgradeableProxy",
        impl_sol_file="SpokeInstance.sol",
        impl_contract_name="SpokeInstance",
    ),
    "AaveOracle": ArtifactInfo("AaveOracle.sol", "AaveOracle"),
    "SignatureGateway": ArtifactInfo("SignatureGateway.sol", "SignatureGateway"),
    "NativeTokenGateway": ArtifactInfo("NativeTokenGateway.sol", "NativeTokenGateway"),
    "GiverPositionManager": ArtifactInfo("GiverPositionManager.sol", "GiverPositionManager"),
    "TakerPositionManager": ArtifactInfo("TakerPositionManager.sol", "TakerPositionManager"),
    "ConfigPositionManager": ArtifactInfo("ConfigPositionManager.sol", "ConfigPositionManager"),
    "TokenizationSpoke": ArtifactInfo(
        "TransparentUpgradeableProxy.sol", "TransparentUpgradeableProxy",
        impl_sol_file="TokenizationSpokeInstance.sol",
        impl_contract_name="TokenizationSpokeInstance",
    ),
    "LiquidationLogic": ArtifactInfo("LiquidationLogic.sol", "LiquidationLogic"),
}


@dataclass(frozen=True)
class ExpectedCompilerSettings:
    solc: str
    optimizer_runs: int
    via_ir: bool
    evm_version: str


DEFAULT_COMPILER_SETTINGS = ExpectedCompilerSettings(
    solc="0.8.28",
    optimizer_runs=44_444_444,
    via_ir=False,
    evm_version="cancun",
)

COMPILER_OVERRIDES: dict[str, ExpectedCompilerSettings] = {
    "HubInstance": ExpectedCompilerSettings(
        solc="0.8.28", optimizer_runs=22_300, via_ir=True, evm_version="cancun",
    ),
    "SpokeInstance": ExpectedCompilerSettings(
        solc="0.8.28", optimizer_runs=750, via_ir=True, evm_version="cancun",
    ),
}


def _expected_compiler(contract_name: str) -> ExpectedCompilerSettings:
    return COMPILER_OVERRIDES.get(contract_name, DEFAULT_COMPILER_SETTINGS)


# ---------------------------------------------------------------------------
# Etherscan API helpers
# ---------------------------------------------------------------------------

_last_etherscan_ts: float = 0.0


def _etherscan_get_source(address: str, api_key: str, chain_id: int = 1) -> dict | None:
    """Query Etherscan ``getsourcecode`` for *address*.

    Applies 220 ms rate-limiting between calls to stay under the free-tier
    5 req/s limit.  Returns the first result dict or ``None`` on failure.
    """
    global _last_etherscan_ts
    elapsed = time.monotonic() - _last_etherscan_ts
    if elapsed < 0.22:
        time.sleep(0.22 - elapsed)

    params = urlencode({
        "chainid": str(chain_id),
        "module": "contract",
        "action": "getsourcecode",
        "address": address,
        "apikey": api_key,
    })
    url = f"https://api.etherscan.io/v2/api?{params}"
    try:
        req = Request(url, headers={"User-Agent": "aave-v4-verify/1.0"})
        with urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode())
    except Exception as e:
        print(f"  {RED}WARNING{RESET} Etherscan request failed for {address}: {e}")
        return None
    finally:
        _last_etherscan_ts = time.monotonic()

    if data.get("status") != "1" or not data.get("result"):
        msg = data.get("message", "unknown error")
        detail = data.get("result", "")
        print(f"  {RED}WARNING{RESET} Etherscan API error for {address}: {msg} — {detail}")
        return None
    return data["result"][0]


@lru_cache(maxsize=None)
def load_deployed_bytecode(
    sol_file: str, contract_name: str
) -> tuple[bytes, dict, list[dict]]:
    """Return (bytecode, immutable_refs, link_refs).

    - immutable_refs: positions of constructor-set immutables (must be masked)
    - link_refs: positions of library address placeholders (will be patched)
    """
    path = ARTIFACTS_DIR / sol_file / f"{contract_name}.json"
    if not path.exists():
        sys.exit(f"Artifact not found: {path}\nRun `forge build` first.")
    with open(path) as f:
        artifact = json.load(f)
    deployed = artifact["deployedBytecode"]
    bytecode_hex = deployed["object"]
    if bytecode_hex.startswith("0x"):
        bytecode_hex = bytecode_hex[2:]

    # Replace unlinked library placeholders (__$<hash>$__) with zero bytes
    # so bytes.fromhex() can parse the string; the positions will be
    # overwritten with real addresses before comparison.
    bytecode_hex = re.sub(r"__\$[0-9a-fA-F]+\$__", "0" * 40, bytecode_hex)

    immutable_refs: dict[str, list[dict]] = {}
    for key, ranges in deployed.get("immutableReferences", {}).items():
        immutable_refs[key] = ranges

    link_refs: list[dict] = []
    for file_refs in deployed.get("linkReferences", {}).values():
        for _, ranges in file_refs.items():
            link_refs.extend(ranges)

    return bytes.fromhex(bytecode_hex), immutable_refs, link_refs


def link_artifact_bytecode(
    artifact_bytes: bytearray,
    onchain_bytes: bytes,
    link_refs: list[dict],
) -> bytearray:
    """Patch library placeholders in *artifact_bytes* with the real addresses
    extracted from *onchain_bytes*, so comparison can be a full 100% match."""
    if not link_refs:
        return artifact_bytes

    # Extract the address from the first link ref position as the canonical one
    first = link_refs[0]
    address = onchain_bytes[first["start"]:first["start"] + first["length"]]

    # Sanity: every link ref position should contain the same address
    for ref in link_refs[1:]:
        chunk = onchain_bytes[ref["start"]:ref["start"] + ref["length"]]
        if chunk != address:
            raise ValueError(
                f"Library address mismatch in on-chain bytecode: "
                f"offset {first['start']} has {address.hex()}, "
                f"offset {ref['start']} has {chunk.hex()}"
            )

    for ref in link_refs:
        start = ref["start"]
        length = ref["length"]
        artifact_bytes[start:start + length] = address

    return artifact_bytes


def mask_bytecode_ranges(bytecode: bytes, refs: dict) -> bytearray:
    masked = bytearray(bytecode)
    for ranges in refs.values():
        for ref in ranges:
            start = ref["start"]
            length = ref["length"]
            for i in range(start, start + length):
                if i < len(masked):
                    masked[i] = 0
    return masked


def _first_diff_offset(a: bytes | bytearray, b: bytes | bytearray) -> int:
    for i in range(min(len(a), len(b))):
        if a[i] != b[i]:
            return i
    return min(len(a), len(b))


GREEN = "\033[92m"
RED = "\033[91m"
YELLOW = "\033[93m"
BOLD = "\033[1m"
RESET = "\033[0m"


class VerificationResult:
    def __init__(self) -> None:
        self.passed = 0
        self.failed = 0

    def ok(self, label: str, details: str = "") -> None:
        self.passed += 1
        suffix = f" ({details})" if details else ""
        print(f"  {GREEN}OK{RESET}  {label}{suffix}")

    def error(self, label: str, expected: Any, actual: Any) -> None:
        self.failed += 1
        print(f"  {RED}ERROR{RESET}  {label}")
        print(f"         expected: {expected}")
        print(f"         actual:   {actual}")

    def section(self, title: str) -> None:
        print(f"\n{BOLD}=== {title} ==={RESET}\n")

    def summary(self) -> int:
        total = self.passed + self.failed
        print(f"\n{BOLD}--- Summary ---{RESET}")
        print(f"  {GREEN}Passed:{RESET} {self.passed}/{total}")
        if self.failed:
            print(f"  {RED}Failed:{RESET} {self.failed}/{total}")
            return 1
        print("  All checks passed.")
        return 0


@dataclass
class HubInfo:
    label: str
    address: str
    ir_strategy: str


@dataclass
class SpokeInfo:
    label: str
    proxy: str
    oracle: str


@dataclass
class DeployReport:
    salt: str
    access_manager: str
    hub_configurator: str
    spoke_configurator: str
    treasury_spoke: str
    hubs: list[HubInfo] = field(default_factory=list)
    spokes: list[SpokeInfo] = field(default_factory=list)
    signature_gateway: Optional[str] = None
    native_token_gateway: Optional[str] = None
    giver_position_manager: Optional[str] = None
    taker_position_manager: Optional[str] = None
    config_position_manager: Optional[str] = None

    @classmethod
    def from_json(cls, data: dict) -> "DeployReport":
        hubs: list[HubInfo] = []
        hub_addrs = data.get("hub", {})
        ir_addrs = data.get("irStrategy", {})
        for key, addr in hub_addrs.items():
            hubs.append(
                HubInfo(
                    label=key,
                    address=addr,
                    ir_strategy=ir_addrs[key],
                )
            )

        spokes: list[SpokeInfo] = []
        spoke_addrs = data.get("spoke", {})
        oracle_addrs = data.get("oracle", {})
        for key, addr in spoke_addrs.items():
            spokes.append(
                SpokeInfo(
                    label=key,
                    proxy=addr,
                    oracle=oracle_addrs[key],
                )
            )

        return cls(
            salt=data.get("salt", ""),
            access_manager=data["accessManager"],
            hub_configurator=data["hubConfigurator"],
            spoke_configurator=data["spokeConfigurator"],
            treasury_spoke=data["treasurySpoke"],
            hubs=hubs,
            spokes=spokes,
            signature_gateway=data.get("signatureGateway"),
            native_token_gateway=data.get("nativeTokenGateway"),
            giver_position_manager=data.get("giverPositionManager"),
            taker_position_manager=data.get("takerPositionManager"),
            config_position_manager=data.get("configPositionManager"),
        )

    def all_addresses(self) -> list[tuple[str, str]]:
        pairs: list[tuple[str, str]] = [
            ("AccessManager", self.access_manager),
            ("HubConfigurator", self.hub_configurator),
            ("SpokeConfigurator", self.spoke_configurator),
            ("TreasurySpoke", self.treasury_spoke),
        ]
        for h in self.hubs:
            pairs.append((f"{h.label}/Hub", h.address))
            pairs.append((f"{h.label}/InterestRateStrategy", h.ir_strategy))
        for s in self.spokes:
            pairs.append((f"{s.label}/Spoke", s.proxy))
            pairs.append((f"{s.label}/AaveOracle", s.oracle))
        for name, addr in [
            ("SignatureGateway", self.signature_gateway),
            ("NativeTokenGateway", self.native_token_gateway),
            ("GiverPositionManager", self.giver_position_manager),
            ("TakerPositionManager", self.taker_position_manager),
            ("ConfigPositionManager", self.config_position_manager),
        ]:
            if addr:
                pairs.append((name, addr))
        return pairs

    def all_report_keys(self) -> list[tuple[str, str]]:
        pairs: list[tuple[str, str]] = [
            ("accessManager", self.access_manager),
            ("hubConfigurator", self.hub_configurator),
            ("spokeConfigurator", self.spoke_configurator),
            ("treasurySpoke", self.treasury_spoke),
        ]
        for h in self.hubs:
            pairs.append((h.label, h.address))
            pairs.append((h.label, h.ir_strategy))
        for s in self.spokes:
            pairs.append((s.label, s.proxy))
            pairs.append((s.label, s.oracle))
        for name, addr in [
            ("signatureGateway", self.signature_gateway),
            ("nativeTokenGateway", self.native_token_gateway),
            ("giverPositionManager", self.giver_position_manager),
            ("takerPositionManager", self.taker_position_manager),
            ("configPositionManager", self.config_position_manager),
        ]:
            if addr:
                pairs.append((name, addr))
        return pairs

    def hub_by_label(self, label: str) -> HubInfo:
        for h in self.hubs:
            if h.label == label:
                return h
        raise KeyError(f"Hub '{label}' not found in report")

    def spoke_by_label(self, label: str) -> SpokeInfo:
        for s in self.spokes:
            if s.label == label:
                return s
        raise KeyError(f"Spoke '{label}' not found in report")


# ---------------------------------------------------------------------------
# XLSX config parser
# ---------------------------------------------------------------------------

_ASSET_NAME_MAP: dict[str, str] = {
    "ETH": "WETH",
    "PT-USDE-7MAY2026": "PT_USDE_7MAY2026",
    "PT-sUSDE-7MAY2026": "PT_sUSDE_7MAY2026",
}

_HUB_NAME_MAP: dict[str, str] = {
    "Core Hub": "CORE_HUB",
    "Prime Hub": "PRIME_HUB",
    "Plus Hub": "PLUS_HUB",
}

_SPOKE_NAME_MAP: dict[str, str] = {
    "Main Spoke": "MAIN_SPOKE",
    "Lido Spoke": "LIDO_ESPOKE",
    "EtherFi Spoke": "ETHERFI_ESPOKE",
    "Kelp Spoke": "KELP_ESPOKE",
    "Gold Spoke": "GOLD_SPOKE",
    "Forex Spoke": "FOREX_SPOKE",
    "Lombard BTC Spoke": "LOMBARD_BTC_SPOKE",
    "Bluechip Spoke": "BLUECHIP_SPOKE",
    "Ethena Ecosystem Spoke": "ETHENA_ECOSYSTEM_SPOKE",
    "Ethena Correlated Spoke": "ETHENA_CORRELATED_SPOKE",
}


def _norm_asset(name: str) -> str:
    return _ASSET_NAME_MAP.get(name, name)


def _norm_hub(name: str) -> str:
    return _HUB_NAME_MAP.get(name, name)


def _norm_spoke(name: str) -> str:
    return _SPOKE_NAME_MAP.get(name, name)


def _is_tokenization_spoke(spoke_name: str) -> bool:
    return "Tokenized" in spoke_name


def _tokenize_name_symbol(hub_key: str, token_key: str) -> tuple[str, str]:
    """Derive TokenizationSpoke name and symbol from hub and token keys."""
    hub_short = {"CORE_HUB": "Core", "PRIME_HUB": "Prime", "PLUS_HUB": "Plus"}[hub_key]
    if token_key.startswith("PT_"):
        return f"Tokenized Aave {hub_short} {token_key}", f"a{hub_short}-{token_key}"
    return f"Wrapped Aave {hub_short} {token_key}", f"wa{hub_short}{token_key}"


def _to_bps(value) -> int:
    """Convert a decimal value to basis points (multiply by 10000)."""
    if isinstance(value, (int, float)):
        return int(round(value * 10000))
    return 0


def _to_wad(value: float) -> str:
    """Convert a decimal to a WAD string (multiply by 1e18) without float precision loss."""
    d = Decimal(str(value)).quantize(Decimal("0.0001"))
    return str(int(d * 10**18))


def _parse_base_rate(value) -> int:
    """Parse base rate: "0" → 0, "0.25%" → 25, numeric → bps."""
    if isinstance(value, (int, float)):
        return _to_bps(value)
    if isinstance(value, str):
        v = value.strip()
        if v == "0":
            return 0
        if v.endswith("%"):
            return int(round(float(v[:-1]) * 100))
    return 0


def _parse_reserve_level_params(ws) -> tuple[list[dict], list[dict], dict[tuple[str, str], int]]:
    """Parse 'Reserve Level Params' sheet.

    Returns (spoke_registrations, reserves, tokenize_caps).
    tokenize_caps maps (hub_key, token_key) → addCap for tokenization spokes.
    """
    spoke_regs: list[dict] = []
    reserves: list[dict] = []
    tokenize_caps: dict[tuple[str, str], int] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if len(vals) < 11 or vals[0] is None:
            continue

        _chain, hub_name, spoke_name, asset_name = vals[0], vals[1], vals[2], vals[3]
        add_cap, draw_cap = vals[4], vals[5]
        cf, mlb, borrowable, cr, lf = vals[6], vals[7], vals[8], vals[9], vals[10]

        if not hub_name or not spoke_name or not asset_name:
            continue

        hub_key = _norm_hub(hub_name)
        token_key = _norm_asset(asset_name)

        # Tokenization spokes → extract addCap, skip from reserves/spoke_regs
        if _is_tokenization_spoke(spoke_name):
            cap = int(add_cap) if isinstance(add_cap, (int, float)) else 0
            tokenize_caps[(hub_key, token_key)] = cap
            continue

        spoke_key = _norm_spoke(spoke_name)

        spoke_regs.append({
            "assetKey": token_key,
            "hubKey": hub_key,
            "spokeKey": spoke_key,
            "addCap": int(add_cap) if isinstance(add_cap, (int, float)) else 0,
            "drawCap": int(draw_cap) if isinstance(draw_cap, (int, float)) else 0,
        })

        reserve = {
            "spokeKey": spoke_key,
            "hubKey": hub_key,
            "assetKey": token_key,
            "borrowable": bool(borrowable),
            "collateralFactor": _to_bps(cf),
            "collateralRisk": _to_bps(cr) if isinstance(cr, (int, float)) else 0,
        }
        if isinstance(mlb, (int, float)):
            reserve["maxLiquidationBonus"] = int(round((1 + mlb) * 10000))
        if isinstance(lf, (int, float)):
            reserve["liquidationFee"] = int(round(lf * 10000))

        reserves.append(reserve)

    return spoke_regs, reserves, tokenize_caps


def _parse_asset_level_ir_params(ws, tokenize_caps: dict[tuple[str, str], int]) -> list[dict]:
    """Parse 'Asset Level IR Params' sheet → assets list."""
    assets: list[dict] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if len(vals) < 8 or vals[0] is None:
            continue

        _chain, hub_name, asset_name = vals[0], vals[1], vals[2]
        base, slope1, slope2, uoptimal, liq_fee = vals[3], vals[4], vals[5], vals[6], vals[7]

        if not hub_name or not asset_name:
            continue

        hub_key = _norm_hub(hub_name)
        token_key = _norm_asset(asset_name)

        entry: dict = {"tokenKey": token_key, "hubKey": hub_key}

        if base == "N/A" or slope1 == "N/A":
            entry["irData"] = {
                "optimalUsageRatio": 9900,
                "baseDrawnRate": 0,
                "rateGrowthBeforeOptimal": 0,
                "rateGrowthAfterOptimal": 0,
            }
        else:
            entry["irData"] = {
                "baseDrawnRate": _parse_base_rate(base),
                "rateGrowthBeforeOptimal": _to_bps(slope1),
                "rateGrowthAfterOptimal": _to_bps(slope2),
                "optimalUsageRatio": _to_bps(uoptimal),
            }
            if isinstance(liq_fee, (int, float)):
                entry["liquidityFee"] = _to_bps(liq_fee)

        # Merge tokenization addCap if available
        cap = tokenize_caps.get((hub_key, token_key))
        if cap is not None:
            name, symbol = _tokenize_name_symbol(hub_key, token_key)
            entry["tokenize"] = {"name": name, "symbol": symbol, "addCap": cap}

        assets.append(entry)

    return assets


def _parse_spoke_level_params(ws) -> list[dict]:
    """Parse 'Spoke Level Params' sheet → spokes list."""
    spokes: list[dict] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if len(vals) < 6 or vals[0] is None:
            continue

        _chain, _hub_name, spoke_name = vals[0], vals[1], vals[2]
        lbf, thf, hfmb = vals[3], vals[4], vals[5]

        if not spoke_name:
            continue

        spoke_key = _norm_spoke(spoke_name)

        spokes.append({
            "key": spoke_key,
            "liquidationConfig": {
                "liquidationBonusFactor": _to_bps(lbf),
                "targetHealthFactor": _to_wad(thf),
                "healthFactorForMaxBonus": _to_wad(hfmb),
            },
        })

    return spokes


def load_config_from_xlsx(path: str) -> dict:
    """Load configuration from an xlsx file and return a dict for ConfigInput."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)

    spoke_regs, reserves, tokenize_caps = _parse_reserve_level_params(
        wb["Reserve Level Params"]
    )
    assets = _parse_asset_level_ir_params(
        wb["Asset Level IR Params"], tokenize_caps
    )
    spokes = _parse_spoke_level_params(wb["Spoke Level Params"])

    # Derive hubs from reserves
    seen_hubs: dict[str, dict] = {}
    for r in reserves:
        hk = r["hubKey"]
        if hk not in seen_hubs:
            seen_hubs[hk] = {"key": hk}

    return {
        "defaults": {
            "spokeRegistration": {
                "riskPremiumThreshold": 0,
                "active": True,
                "halted": False,
            },
            "reserve": {
                "receiveSharesEnabled": True,
                "frozen": False,
                "paused": False,
            },
        },
        "tokens": {},
        "hubs": list(seen_hubs.values()),
        "spokes": spokes,
        "assets": assets,
        "spokeRegistrations": spoke_regs,
        "reserves": reserves,
    }


class ConfigInput:
    def __init__(self, data: dict) -> None:
        self._raw = data
        self.defaults: dict = data.get("defaults", {})
        self.tokens_by_key: dict[str, dict] = data.get("tokens", {})
        self.hubs_by_key: dict[str, dict] = {
            h["key"]: h for h in data.get("hubs", [])
        }
        self.spokes_by_key: dict[str, dict] = {
            s["key"]: s for s in data.get("spokes", [])
        }
        self.assets: list[dict] = data.get("assets", [])
        self.spoke_registrations: list[dict] = data.get("spokeRegistrations", [])
        self.reserves: list[dict] = data.get("reserves", [])

    def resolve_default(
        self, entry: dict, field_name: str, defaults_section: str
    ) -> Any:
        if field_name in entry:
            return entry[field_name]
        section = self.defaults.get(defaults_section, {})
        return section[field_name]


class ContractCaller:
    def __init__(self, w3: Web3) -> None:
        self.w3 = w3
        self.hub_abi = load_abi("IHub.sol", "IHub")
        self.spoke_abi = load_abi("ISpoke.sol", "ISpoke")
        self.ir_strategy_abi = load_abi("IAssetInterestRateStrategy.sol", "IAssetInterestRateStrategy")
        self.oracle_abi = load_abi("IAaveOracle.sol", "IAaveOracle")
        self.tokenization_spoke_abi = load_abi("ITokenizationSpoke.sol", "ITokenizationSpoke")
        self.access_manager_abi = load_abi(
            "IAccessManagerEnumerable.sol", "IAccessManagerEnumerable"
        )
        self.position_manager_base_abi = load_abi(
            "IPositionManagerBase.sol", "IPositionManagerBase"
        )
        self.price_oracle_abi = load_abi("IPriceOracle.sol", "IPriceOracle")
        self.price_feed_abi = load_abi("IPriceFeed.sol", "IPriceFeed")

    ERC1967_IMPL_SLOT = int(
        "0x360894a13ba1a3210667c828492db98dca3e2076cc3735a920a3ca505d382bbc", 16
    )

    def get_code(self, address: str) -> bytes:
        return self.w3.eth.get_code(Web3.to_checksum_address(address))

    def get_implementation(self, proxy_address: str) -> str:
        raw = self.w3.eth.get_storage_at(
            Web3.to_checksum_address(proxy_address), self.ERC1967_IMPL_SLOT
        )
        return Web3.to_checksum_address("0x" + raw[-20:].hex())

    def _call(self, address: str, abi: list, fn_name: str, *args: Any, silent: bool = False) -> Any:
        try:
            contract = self.w3.eth.contract(
                address=Web3.to_checksum_address(address), abi=abi
            )
            return contract.functions[fn_name](*args).call()
        except Exception as e:
            if not silent:
                print(f"  {RED}CALL FAILED{RESET} {fn_name}({', '.join(str(a) for a in args)}) on {address}: {e}")
            return None

    def call_hub(self, hub_addr: str, fn_name: str, *args: Any) -> Any:
        return self._call(hub_addr, self.hub_abi, fn_name, *args)

    def call_spoke(self, spoke_addr: str, fn_name: str, *args: Any) -> Any:
        return self._call(spoke_addr, self.spoke_abi, fn_name, *args)

    def call_ir_strategy(self, addr: str, fn_name: str, *args: Any) -> Any:
        return self._call(addr, self.ir_strategy_abi, fn_name, *args)

    def call_oracle(self, addr: str, fn_name: str, *args: Any) -> Any:
        return self._call(addr, self.oracle_abi, fn_name, *args)

    def call_tokenization_spoke(self, addr: str, fn_name: str, *args: Any, silent: bool = False) -> Any:
        return self._call(addr, self.tokenization_spoke_abi, fn_name, *args, silent=silent)

    def call_access_manager(self, addr: str, fn_name: str, *args: Any) -> Any:
        return self._call(addr, self.access_manager_abi, fn_name, *args)

    def call_price_feed(self, addr: str, fn_name: str, *args: Any) -> Any:
        return self._call(addr, self.price_feed_abi, fn_name, *args)


# ---------------------------------------------------------------------------
# Batch RPC infrastructure
# ---------------------------------------------------------------------------

MAX_BATCH_SIZE = 100

class BatchCallManager:
    """Wraps web3.py batch_requests() to execute many eth_call requests in one
    JSON-RPC batch.  Accepts a list of (key, contract_call_builder) tuples,
    executes them in chunks of *max_batch_size*, and returns a dict keyed by
    the caller-provided keys."""

    def __init__(self, w3: Web3, max_batch_size: int = MAX_BATCH_SIZE) -> None:
        self.w3 = w3
        self.max_batch_size = max_batch_size

    def execute(
        self, calls: list[tuple[str, Any]]
    ) -> tuple[dict[str, Any], dict[str, str]]:
        """Execute *calls* — a list of ``(key, contract_fn_call)`` where
        ``contract_fn_call`` is e.g.
        ``contract.functions.getAssetId(token_addr)`` (no ``.call()``).

        Returns ``(results, errors)`` dicts keyed by *key*.
        """
        results: dict[str, Any] = {}
        errors: dict[str, str] = {}
        if not calls:
            return results, errors

        # Process in chunks to respect node batch-size limits
        for chunk_start in range(0, len(calls), self.max_batch_size):
            chunk = calls[chunk_start : chunk_start + self.max_batch_size]
            chunk_keys = [k for k, _ in chunk]

            batch = self.w3.batch_requests()
            for _, fn_call in chunk:
                batch.add(fn_call)

            try:
                responses = batch.execute()
            except Exception as e:
                for key in chunk_keys:
                    errors[key] = str(e)
                continue

            for i, key in enumerate(chunk_keys):
                try:
                    results[key] = responses[i]
                except Exception as e:
                    errors[key] = str(e)

        return results, errors


@dataclass
class DeploymentCache:
    """Pre-fetched shared data used across all verify functions."""
    # (hub_addr, token_addr) -> asset_id or None
    asset_ids: dict[tuple[str, str], int | None] = field(default_factory=dict)
    # (spoke_addr, hub_addr, asset_id) -> reserve_id or None
    reserve_ids: dict[tuple[str, str, int], int | None] = field(default_factory=dict)


def prefetch_shared_data(
    batch_mgr: BatchCallManager,
    caller: ContractCaller,
    report: DeployReport,
    config: ConfigInput,
) -> DeploymentCache:
    """Pre-fetch all asset_ids and reserve_ids in 2 batch round-trips."""
    cache = DeploymentCache()

    if not config.tokens_by_key:
        print(f"  {YELLOW}SKIPPED{RESET} prefetch — no token addresses in config")
        return cache

    # Collect all unique (hub_addr, token_addr) pairs
    hub_token_pairs: dict[tuple[str, str], str] = {}  # -> key for dedup
    for entry in config.assets:
        token_info = config.tokens_by_key.get(entry["tokenKey"])
        if not token_info or "address" not in token_info:
            continue
        hub_info = report.hub_by_label(entry["hubKey"])
        token_addr = Web3.to_checksum_address(token_info["address"])
        hub_addr = Web3.to_checksum_address(hub_info.address)
        hub_token_pairs[(hub_addr, token_addr)] = f"{hub_addr}:{token_addr}"
    for entry in config.spoke_registrations:
        token_info = config.tokens_by_key.get(entry["assetKey"])
        if not token_info or "address" not in token_info:
            continue
        hub_info = report.hub_by_label(entry["hubKey"])
        token_addr = Web3.to_checksum_address(token_info["address"])
        hub_addr = Web3.to_checksum_address(hub_info.address)
        hub_token_pairs[(hub_addr, token_addr)] = f"{hub_addr}:{token_addr}"
    for entry in config.reserves:
        token_info = config.tokens_by_key.get(entry["assetKey"])
        if not token_info or "address" not in token_info:
            continue
        hub_info = report.hub_by_label(entry["hubKey"])
        token_addr = Web3.to_checksum_address(token_info["address"])
        hub_addr = Web3.to_checksum_address(hub_info.address)
        hub_token_pairs[(hub_addr, token_addr)] = f"{hub_addr}:{token_addr}"

    # Round 1: batch all getAssetId calls
    asset_id_calls: list[tuple[str, Any]] = []
    for (hub_addr, token_addr), key in hub_token_pairs.items():
        contract = batch_mgr.w3.eth.contract(
            address=hub_addr, abi=caller.hub_abi
        )
        asset_id_calls.append((key, contract.functions.getAssetId(token_addr)))

    results, errors = batch_mgr.execute(asset_id_calls)
    for (hub_addr, token_addr), key in hub_token_pairs.items():
        if key in errors:
            print(f"  {RED}CALL FAILED{RESET} getAssetId prefetch for {key}: {errors[key]}")
            cache.asset_ids[(hub_addr, token_addr)] = None
        else:
            cache.asset_ids[(hub_addr, token_addr)] = results.get(key)

    # Collect all unique (spoke_addr, hub_addr, asset_id) for reserves + oracles
    reserve_id_keys: dict[tuple[str, str, int], str] = {}
    for entry in config.reserves:
        token_info = config.tokens_by_key.get(entry["assetKey"])
        if not token_info or "address" not in token_info:
            continue
        hub_info = report.hub_by_label(entry["hubKey"])
        spoke_info = report.spoke_by_label(entry["spokeKey"])
        token_addr = Web3.to_checksum_address(token_info["address"])
        hub_addr = Web3.to_checksum_address(hub_info.address)
        spoke_addr = Web3.to_checksum_address(spoke_info.proxy)
        asset_id = cache.asset_ids.get((hub_addr, token_addr))
        if asset_id is None:
            continue
        triple = (spoke_addr, hub_addr, asset_id)
        reserve_id_keys[triple] = f"{spoke_addr}:{hub_addr}:{asset_id}"

    # Round 2: batch all getReserveId calls
    reserve_id_calls: list[tuple[str, Any]] = []
    for (spoke_addr, hub_addr, asset_id), key in reserve_id_keys.items():
        contract = batch_mgr.w3.eth.contract(
            address=spoke_addr, abi=caller.spoke_abi
        )
        reserve_id_calls.append((key, contract.functions.getReserveId(hub_addr, asset_id)))

    results, errors = batch_mgr.execute(reserve_id_calls)
    for (spoke_addr, hub_addr, asset_id), key in reserve_id_keys.items():
        if key in errors:
            print(f"  {RED}CALL FAILED{RESET} getReserveId prefetch for {key}: {errors[key]}")
            cache.reserve_ids[(spoke_addr, hub_addr, asset_id)] = None
        else:
            cache.reserve_ids[(spoke_addr, hub_addr, asset_id)] = results.get(key)

    return cache


# ---------------------------------------------------------------------------
# Verification functions
# ---------------------------------------------------------------------------


def _verify_single_bytecode(
    caller: ContractCaller,
    result: VerificationResult,
    label: str,
    address: str,
    sol_file: str,
    contract_name: str,
) -> None:
    expected_bytes, immutable_refs, link_refs = load_deployed_bytecode(
        sol_file, contract_name
    )
    onchain_code = caller.get_code(address)

    if not onchain_code or onchain_code in (b"", b"\x00"):
        result.error(label, "deployed bytecode", f"empty code at {address}")
        return

    if len(expected_bytes) != len(onchain_code):
        result.error(
            label,
            f"bytecode length {len(expected_bytes)}",
            f"bytecode length {len(onchain_code)} at {address}",
        )
        return

    expected = bytearray(expected_bytes)
    onchain = bytes(onchain_code)

    # Bytecode verification uses up to three steps depending on the contract:
    #
    # 1. Library linking (link_refs): the compiled artifact contains placeholder
    #    slots where external library addresses will be inserted by the linker.
    #    We read the real addresses from the on-chain bytecode at those offsets
    #    and patch them into the artifact so the two can be compared directly.
    #    Example: SpokeInstance links against deployed library contracts.
    #
    # 2. Immutable masking (immutable_refs): values set in the constructor
    #    (e.g. a proxy admin address) are baked into deployed bytecode by the
    #    compiler but are absent from the compiled artifact. Since these bytes
    #    will always differ, we zero them out on both sides before comparing.
    #    Example: proxy contracts whose admin is a constructor argument.
    #
    # 3. Direct match: contracts with no link_refs and no immutable_refs
    #    (e.g. Hub implementations) are compared byte-for-byte with no masking.
    #
    # When masking is applied the result is tagged "(masked)" to signal that
    # immutable positions were excluded — a slightly weaker guarantee than a
    # full direct match, but expected for contracts with constructor immutables.

    # Patch library placeholders with real addresses from on-chain bytecode
    if link_refs:
        expected = link_artifact_bytecode(expected, onchain, link_refs)

    # Mask immutable positions (constructor-set values differ per deployment)
    if immutable_refs:
        expected_cmp = mask_bytecode_ranges(bytes(expected), immutable_refs)
        onchain_cmp = mask_bytecode_ranges(onchain, immutable_refs)
        tag = "masked"
    else:
        expected_cmp = expected
        onchain_cmp = bytearray(onchain)
        tag = ""

    if expected_cmp == onchain_cmp:
        suffix = f"{address} ({tag})" if tag else address
        result.ok(label, suffix)
    else:
        offset = _first_diff_offset(expected_cmp, onchain_cmp)
        kind = f"matching bytecode ({tag})" if tag else "matching bytecode"
        result.error(
            label,
            kind,
            f"mismatch at byte offset {offset} at {address}",
        )


def verify_liquidation_logic_libraries(
    caller: ContractCaller, report: DeployReport, result: VerificationResult
) -> None:
    """Verify LiquidationLogic library address and bytecode for each Spoke."""
    artifact = ARTIFACT_MAP["LiquidationLogic"]
    addresses: dict[str, str] = {}  # spoke_label -> lib address

    for spoke in report.spokes:
        spoke_addr = Web3.to_checksum_address(spoke.proxy)
        lib_addr = caller.call_spoke(spoke_addr, "getLiquidationLogic")
        label = f"{spoke.label}/LiquidationLogic"

        if lib_addr is None or lib_addr == "0x" + "0" * 40:
            result.error(label, "non-zero library address", str(lib_addr))
            continue

        lib_addr = Web3.to_checksum_address(lib_addr)
        addresses[spoke.label] = lib_addr

        _verify_single_bytecode(
            caller, result, label, lib_addr,
            artifact.sol_file, artifact.contract_name,
        )

    unique_addrs = set(addresses.values())
    if len(unique_addrs) == 1:
        result.ok("LiquidationLogic/consistency", f"all spokes use {unique_addrs.pop()}")
    elif len(unique_addrs) > 1:
        details = ", ".join(f"{lbl}={addr}" for lbl, addr in addresses.items())
        result.error("LiquidationLogic/consistency", "same address across all spokes", details)


def verify_bytecode(
    caller: ContractCaller, report: DeployReport, result: VerificationResult
) -> None:
    result.section("Bytecode Verification")
    for name, addr in report.all_addresses():
        suffix = name.rsplit("/", 1)[-1]
        artifact = ARTIFACT_MAP.get(suffix)
        if artifact is None:
            result.error(name, "known artifact mapping", f"no mapping for '{suffix}'")
            continue

        _verify_single_bytecode(
            caller, result, name, addr,
            artifact.sol_file, artifact.contract_name,
        )

        if artifact.impl_sol_file:
            impl_addr = caller.get_implementation(addr)
            _verify_single_bytecode(
                caller, result, f"{name}/Implementation", impl_addr,
                artifact.impl_sol_file, artifact.impl_contract_name,
            )

    verify_liquidation_logic_libraries(caller, report, result)


def verify_hub_assets(
    batch_mgr: BatchCallManager,
    caller: ContractCaller,
    cache: DeploymentCache,
    report: DeployReport,
    config: ConfigInput,
    result: VerificationResult,
) -> None:
    result.section("Hub Assets & Interest Rate Configuration")

    if not config.tokens_by_key:
        print(f"  {YELLOW}SKIPPED{RESET} — no token addresses in config")
        return

    # Build batch: getInterestRateData + getAssetConfig for each asset
    ir_calls: list[tuple[str, Any]] = []
    cfg_calls: list[tuple[str, Any]] = []
    valid_entries: list[tuple[str, dict, HubInfo, int]] = []

    for entry in config.assets:
        hub_key = entry["hubKey"]
        token_key = entry["tokenKey"]
        token_info = config.tokens_by_key.get(token_key)
        if not token_info or "address" not in token_info:
            continue
        hub_info = report.hub_by_label(hub_key)
        token_addr = Web3.to_checksum_address(token_info["address"])
        hub_addr = Web3.to_checksum_address(hub_info.address)
        label = f"{hub_key}/{token_key}"

        asset_id = cache.asset_ids.get((hub_addr, token_addr))
        if asset_id is None:
            result.error(label, "valid assetId", "call failed")
            continue

        valid_entries.append((label, entry, hub_info, asset_id))

        ir_contract = batch_mgr.w3.eth.contract(
            address=Web3.to_checksum_address(hub_info.ir_strategy),
            abi=caller.ir_strategy_abi,
        )
        ir_calls.append((f"{label}/IR", ir_contract.functions.getInterestRateData(asset_id)))

        hub_contract = batch_mgr.w3.eth.contract(
            address=hub_addr, abi=caller.hub_abi,
        )
        cfg_calls.append((f"{label}/cfg", hub_contract.functions.getAssetConfig(asset_id)))

    ir_results, ir_errors = batch_mgr.execute(ir_calls)
    cfg_results, cfg_errors = batch_mgr.execute(cfg_calls)

    for label, entry, hub_info, asset_id in valid_entries:
        ir_key = f"{label}/IR"
        if ir_key in ir_errors:
            result.error(ir_key, "interest rate data", f"call failed: {ir_errors[ir_key]}")
        else:
            ir_data = ir_results.get(ir_key)
            if ir_data is None:
                result.error(ir_key, "interest rate data", "call failed")
            elif "irData" in entry:
                ir_input = entry["irData"]
                _check(result, f"{label}/optimalUsageRatio", ir_input["optimalUsageRatio"], ir_data[0])
                _check(result, f"{label}/baseDrawnRate", ir_input["baseDrawnRate"], ir_data[1])
                _check(result, f"{label}/rateGrowthBeforeOptimal", ir_input["rateGrowthBeforeOptimal"], ir_data[2])
                _check(result, f"{label}/rateGrowthAfterOptimal", ir_input["rateGrowthAfterOptimal"], ir_data[3])

        cfg_key = f"{label}/cfg"
        if cfg_key in cfg_errors:
            result.error(f"{label}/assetConfig", "asset config", f"call failed: {cfg_errors[cfg_key]}")
        else:
            asset_cfg = cfg_results.get(cfg_key)
            if asset_cfg is None:
                result.error(f"{label}/assetConfig", "asset config", "call failed")
            else:
                expected_fee = entry.get(
                    "liquidityFee",
                    config.defaults.get("asset", {}).get("liquidityFee", 0),
                )
                _check(result, f"{label}/liquidityFee", expected_fee, asset_cfg[1])


def verify_spoke_registrations(
    batch_mgr: BatchCallManager,
    caller: ContractCaller,
    cache: DeploymentCache,
    report: DeployReport,
    config: ConfigInput,
    result: VerificationResult,
) -> None:
    result.section("Spoke Registrations (Hub-side)")

    if not config.spoke_registrations:
        print(f"  {YELLOW}SKIPPED{RESET} — no spoke registration data in config")
        return

    spoke_cfg_calls: list[tuple[str, Any]] = []
    valid_entries: list[tuple[str, dict]] = []

    for entry in config.spoke_registrations:
        hub_key = entry["hubKey"]
        spoke_key = entry["spokeKey"]
        asset_key = entry["assetKey"]
        hub_info = report.hub_by_label(hub_key)
        spoke_info = report.spoke_by_label(spoke_key)
        token_info = config.tokens_by_key.get(asset_key)
        if not token_info or "address" not in token_info:
            continue
        token_addr = Web3.to_checksum_address(token_info["address"])
        hub_addr = Web3.to_checksum_address(hub_info.address)
        label = f"{hub_key}/{spoke_key}/{asset_key}"

        asset_id = cache.asset_ids.get((hub_addr, token_addr))
        if asset_id is None:
            result.error(label, "valid assetId", "call failed")
            continue

        spoke_proxy = Web3.to_checksum_address(spoke_info.proxy)
        hub_contract = batch_mgr.w3.eth.contract(
            address=hub_addr, abi=caller.hub_abi,
        )
        spoke_cfg_calls.append((label, hub_contract.functions.getSpokeConfig(asset_id, spoke_proxy)))
        valid_entries.append((label, entry))

    results, errors = batch_mgr.execute(spoke_cfg_calls)

    for label, entry in valid_entries:
        if label in errors:
            result.error(f"{label}/spokeConfig", "spoke config", f"call failed: {errors[label]}")
            continue
        spoke_cfg = results.get(label)
        if spoke_cfg is None:
            result.error(f"{label}/spokeConfig", "spoke config", "call failed")
            continue

        _check(result, f"{label}/addCap", entry["addCap"], spoke_cfg[0])
        _check(result, f"{label}/drawCap", entry["drawCap"], spoke_cfg[1])
        _check(
            result,
            f"{label}/riskPremiumThreshold",
            config.resolve_default(entry, "riskPremiumThreshold", "spokeRegistration"),
            spoke_cfg[2],
        )
        _check(
            result,
            f"{label}/active",
            config.resolve_default(entry, "active", "spokeRegistration"),
            spoke_cfg[3],
        )
        _check(
            result,
            f"{label}/halted",
            config.resolve_default(entry, "halted", "spokeRegistration"),
            spoke_cfg[4],
        )


def verify_reserves(
    batch_mgr: BatchCallManager,
    caller: ContractCaller,
    cache: DeploymentCache,
    report: DeployReport,
    config: ConfigInput,
    result: VerificationResult,
) -> None:
    result.section("Reserve Configuration (Spoke-side)")

    if not config.tokens_by_key:
        print(f"  {YELLOW}SKIPPED{RESET} — no token addresses in config")
        return

    # Batch A: getReserveConfig + getReserve for each reserve
    rcfg_calls: list[tuple[str, Any]] = []
    reserve_calls: list[tuple[str, Any]] = []
    valid_entries: list[tuple[str, dict, str, int]] = []  # label, entry, spoke_addr, reserve_id

    for entry in config.reserves:
        hub_key = entry["hubKey"]
        spoke_key = entry["spokeKey"]
        asset_key = entry["assetKey"]
        hub_info = report.hub_by_label(hub_key)
        spoke_info = report.spoke_by_label(spoke_key)
        token_info = config.tokens_by_key.get(asset_key)
        if not token_info or "address" not in token_info:
            continue
        token_addr = Web3.to_checksum_address(token_info["address"])
        hub_addr = Web3.to_checksum_address(hub_info.address)
        spoke_addr = Web3.to_checksum_address(spoke_info.proxy)
        label = f"{spoke_key}/{hub_key}/{asset_key}"

        asset_id = cache.asset_ids.get((hub_addr, token_addr))
        if asset_id is None:
            result.error(label, "valid assetId", "call failed")
            continue

        reserve_id = cache.reserve_ids.get((spoke_addr, hub_addr, asset_id))
        if reserve_id is None:
            result.error(f"{label}/reserveId", "valid reserveId", "call failed")
            continue

        valid_entries.append((label, entry, spoke_addr, reserve_id))

        spoke_contract = batch_mgr.w3.eth.contract(
            address=spoke_addr, abi=caller.spoke_abi,
        )
        rcfg_calls.append((f"{label}/rcfg", spoke_contract.functions.getReserveConfig(reserve_id)))
        reserve_calls.append((f"{label}/reserve", spoke_contract.functions.getReserve(reserve_id)))

    rcfg_results, rcfg_errors = batch_mgr.execute(rcfg_calls)
    reserve_results, reserve_errors = batch_mgr.execute(reserve_calls)

    # Check ReserveConfig results and build Batch B for getDynamicReserveConfig
    drc_calls: list[tuple[str, Any]] = []
    drc_entries: list[tuple[str, dict]] = []

    for label, entry, spoke_addr, reserve_id in valid_entries:
        rcfg_key = f"{label}/rcfg"
        if rcfg_key in rcfg_errors:
            result.error(f"{label}/reserveConfig", "reserve config", f"call failed: {rcfg_errors[rcfg_key]}")
        else:
            rcfg = rcfg_results.get(rcfg_key)
            if rcfg is None:
                result.error(f"{label}/reserveConfig", "reserve config", "call failed")
            else:
                if "collateralRisk" in entry:
                    _check(result, f"{label}/collateralRisk", entry["collateralRisk"], rcfg[0])
                if config.defaults.get("reserve") or "paused" in entry:
                    _check(
                        result,
                        f"{label}/paused",
                        config.resolve_default(entry, "paused", "reserve"),
                        rcfg[1],
                    )
                if config.defaults.get("reserve") or "frozen" in entry:
                    _check(
                        result,
                        f"{label}/frozen",
                        config.resolve_default(entry, "frozen", "reserve"),
                        rcfg[2],
                    )
                _check(result, f"{label}/borrowable", entry["borrowable"], rcfg[3])
                if config.defaults.get("reserve") or "receiveSharesEnabled" in entry:
                    _check(
                        result,
                        f"{label}/receiveSharesEnabled",
                        config.resolve_default(entry, "receiveSharesEnabled", "reserve"),
                        rcfg[4],
                )

        reserve_key = f"{label}/reserve"
        if reserve_key in reserve_errors:
            result.error(f"{label}/reserve", "reserve data", f"call failed: {reserve_errors[reserve_key]}")
            continue
        reserve = reserve_results.get(reserve_key)
        if reserve is None:
            result.error(f"{label}/reserve", "reserve data", "call failed")
            continue

        dynamic_config_key = reserve[6]
        spoke_contract = batch_mgr.w3.eth.contract(
            address=spoke_addr, abi=caller.spoke_abi,
        )
        drc_calls.append((f"{label}/drc", spoke_contract.functions.getDynamicReserveConfig(reserve_id, dynamic_config_key)))
        drc_entries.append((label, entry))

    # Batch B: getDynamicReserveConfig
    drc_results, drc_errors = batch_mgr.execute(drc_calls)

    for label, entry in drc_entries:
        drc_key = f"{label}/drc"
        if drc_key in drc_errors:
            result.error(f"{label}/dynamicReserveConfig", "dynamic config", f"call failed: {drc_errors[drc_key]}")
        else:
            drc = drc_results.get(drc_key)
            if drc is None:
                result.error(f"{label}/dynamicReserveConfig", "dynamic config", "call failed")
            else:
                cf = entry.get("collateralFactor")
                if cf is not None:
                    _check(result, f"{label}/collateralFactor", cf, drc[0])
                if "maxLiquidationBonus" in entry or "maxLiquidationBonus" in config.defaults.get("reserve", {}):
                    _check(
                        result,
                        f"{label}/maxLiquidationBonus",
                        config.resolve_default(entry, "maxLiquidationBonus", "reserve"),
                        drc[1],
                    )
                if "liquidationFee" in entry or "liquidationFee" in config.defaults.get("reserve", {}):
                    _check(
                        result,
                        f"{label}/liquidationFee",
                        config.resolve_default(entry, "liquidationFee", "reserve"),
                        drc[2],
                    )


def verify_liquidation_configs(
    batch_mgr: BatchCallManager,
    caller: ContractCaller,
    report: DeployReport,
    config: ConfigInput,
    result: VerificationResult,
) -> None:
    result.section("Liquidation Configuration")

    has_liq_config = any(
        "liquidationConfig" in s for s in config.spokes_by_key.values()
    )
    if not has_liq_config:
        print(f"  {YELLOW}SKIPPED{RESET} — no liquidation config data in config")
        return

    calls: list[tuple[str, Any]] = []
    entries: list[tuple[str, dict]] = []

    for spoke_entry in config.spokes_by_key.values():
        spoke_key = spoke_entry["key"]
        spoke_info = report.spoke_by_label(spoke_key)
        spoke_contract = batch_mgr.w3.eth.contract(
            address=Web3.to_checksum_address(spoke_info.proxy),
            abi=caller.spoke_abi,
        )
        calls.append((spoke_key, spoke_contract.functions.getLiquidationConfig()))
        entries.append((spoke_key, spoke_entry))

    results, errors = batch_mgr.execute(calls)

    for label, spoke_entry in entries:
        if label in errors:
            result.error(f"{label}/liquidationConfig", "liquidation config", f"call failed: {errors[label]}")
            continue
        liq_cfg = results.get(label)
        if liq_cfg is None:
            result.error(f"{label}/liquidationConfig", "liquidation config", "call failed")
            continue

        liq_input = spoke_entry.get("liquidationConfig", {})
        defaults_liq = config.defaults.get("spoke", {}).get("liquidationConfig", {})

        expected_thf = liq_input.get(
            "targetHealthFactor", defaults_liq.get("targetHealthFactor")
        )
        expected_hfmb = liq_input.get(
            "healthFactorForMaxBonus", defaults_liq.get("healthFactorForMaxBonus")
        )
        expected_lbf = liq_input.get(
            "liquidationBonusFactor", defaults_liq.get("liquidationBonusFactor")
        )

        _check(result, f"{label}/targetHealthFactor", expected_thf, liq_cfg[0])
        _check(result, f"{label}/healthFactorForMaxBonus", expected_hfmb, liq_cfg[1])
        _check(result, f"{label}/liquidationBonusFactor", expected_lbf, liq_cfg[2])


def verify_oracles(
    batch_mgr: BatchCallManager,
    caller: ContractCaller,
    cache: DeploymentCache,
    report: DeployReport,
    config: ConfigInput,
    result: VerificationResult,
) -> None:
    result.section("Oracle Configuration")

    if not config.tokens_by_key:
        print(f"  {YELLOW}SKIPPED{RESET} — no token addresses in config")
        return

    expected_decimals = config.defaults.get("spoke", {}).get("oracleDecimals", 8)

    # Batch all decimals() + getReserveSource() calls
    calls: list[tuple[str, Any]] = []

    # Track decimals calls for output
    decimals_keys: list[tuple[str, str]] = []  # (key, label)
    # Track source calls
    source_entries: list[tuple[str, str, str]] = []  # (key, res_label, asset_key)

    for spoke_info in report.spokes:
        oracle_addr = Web3.to_checksum_address(spoke_info.oracle)
        label = spoke_info.label

        oracle_contract = batch_mgr.w3.eth.contract(
            address=oracle_addr, abi=caller.oracle_abi,
        )
        dec_key = f"{label}/oracle/decimals"
        calls.append((dec_key, oracle_contract.functions.decimals()))
        decimals_keys.append((dec_key, label))

        for entry in config.reserves:
            if entry["spokeKey"] != spoke_info.label:
                continue

            hub_key = entry["hubKey"]
            asset_key = entry["assetKey"]
            token_info = config.tokens_by_key.get(asset_key)
            if not token_info or "address" not in token_info:
                continue
            hub_info = report.hub_by_label(hub_key)
            token_addr = Web3.to_checksum_address(token_info["address"])
            hub_addr = Web3.to_checksum_address(hub_info.address)
            spoke_addr = Web3.to_checksum_address(spoke_info.proxy)
            res_label = f"{label}/{hub_key}/{asset_key}"

            asset_id = cache.asset_ids.get((hub_addr, token_addr))
            if asset_id is None:
                result.error(f"{res_label}/oracle/source", "valid assetId", "call failed")
                continue

            reserve_id = cache.reserve_ids.get((spoke_addr, hub_addr, asset_id))
            if reserve_id is None:
                result.error(f"{res_label}/oracle/source", "valid reserveId", "call failed")
                continue

            src_key = f"{res_label}/oracle/source"
            calls.append((src_key, oracle_contract.functions.getReserveSource(reserve_id)))
            source_entries.append((src_key, res_label, asset_key))

    results, errors = batch_mgr.execute(calls)

    for dec_key, label in decimals_keys:
        if dec_key in errors:
            result.error(f"{label}/oracle/decimals", "decimals", f"call failed: {errors[dec_key]}")
        else:
            decimals = results.get(dec_key)
            if decimals is None:
                result.error(f"{label}/oracle/decimals", "decimals", "call failed")
            else:
                _check(result, f"{label}/oracle/decimals", expected_decimals, decimals)

    for src_key, res_label, asset_key in source_entries:
        if src_key in errors:
            result.error(f"{res_label}/oracle/source", "source address", f"call failed: {errors[src_key]}")
        else:
            source = results.get(src_key)
            if source is None:
                result.error(f"{res_label}/oracle/source", "source address", "call failed")
            else:
                token_info = config.tokens_by_key.get(asset_key, {})
                price_feed = token_info.get("priceFeed")
                if price_feed:
                    expected_feed = Web3.to_checksum_address(price_feed)
                    _check(result, f"{res_label}/oracle/source", expected_feed, source)


def verify_oracle_wiring(
    batch_mgr: BatchCallManager,
    caller: ContractCaller,
    report: DeployReport,
    result: VerificationResult,
) -> None:
    result.section("Oracle Wiring Verification")

    calls: list[tuple[str, Any]] = []
    expectations: list[tuple[str, str, str]] = []  # (key, direction, expected_addr)

    for spoke_info in report.spokes:
        spoke_addr = Web3.to_checksum_address(spoke_info.proxy)
        oracle_addr = Web3.to_checksum_address(spoke_info.oracle)

        # Spoke -> Oracle link
        spoke_contract = batch_mgr.w3.eth.contract(
            address=spoke_addr, abi=caller.spoke_abi,
        )
        fwd_key = f"{spoke_info.label}/spoke->oracle"
        calls.append((fwd_key, spoke_contract.functions.ORACLE()))
        expectations.append((fwd_key, "ORACLE()", oracle_addr))

        # Oracle -> Spoke link
        oracle_contract = batch_mgr.w3.eth.contract(
            address=oracle_addr, abi=caller.price_oracle_abi,
        )
        rev_key = f"{spoke_info.label}/oracle->spoke"
        calls.append((rev_key, oracle_contract.functions.spoke()))
        expectations.append((rev_key, "spoke()", spoke_addr))

    results, errors = batch_mgr.execute(calls)

    for key, fn_name, expected_addr in expectations:
        if key in errors:
            result.error(key, expected_addr, f"call failed: {errors[key]}")
        else:
            actual = results.get(key)
            if actual is not None:
                actual = Web3.to_checksum_address(actual)
            _check(result, key, expected_addr, actual)


def verify_price_feeds(
    batch_mgr: BatchCallManager,
    caller: ContractCaller,
    report: DeployReport,
    config: ConfigInput,
    result: VerificationResult,
) -> None:
    result.section("Price Feed Verification")

    if not config.tokens_by_key:
        print(f"  {YELLOW}SKIPPED{RESET} — no token/price feed data in config")
        return

    expected_decimals = config.defaults.get("spoke", {}).get("oracleDecimals", 8)

    # Collect unique price feed addresses across all reserves
    feed_addresses: dict[str, list[str]] = {}  # feed_addr -> [asset_keys that use it]
    for entry in config.reserves:
        asset_key = entry["assetKey"]
        token_info = config.tokens_by_key.get(asset_key, {})
        feed_addr_raw = token_info.get("priceFeed")
        if not feed_addr_raw:
            continue
        feed_addr = Web3.to_checksum_address(feed_addr_raw)
        feed_addresses.setdefault(feed_addr, []).append(asset_key)

    if not feed_addresses:
        print("  No price feeds found in config.")
        return

    calls: list[tuple[str, Any]] = []
    for feed_addr in feed_addresses:
        feed_contract = batch_mgr.w3.eth.contract(
            address=feed_addr, abi=caller.price_feed_abi,
        )
        calls.append((f"{feed_addr}/decimals", feed_contract.functions.decimals()))
        calls.append((f"{feed_addr}/latestAnswer", feed_contract.functions.latestAnswer()))

    results, errors = batch_mgr.execute(calls)

    for feed_addr, asset_keys in feed_addresses.items():
        assets_str = ", ".join(asset_keys)

        dec_key = f"{feed_addr}/decimals"
        if dec_key in errors:
            result.error(f"priceFeed({assets_str})/decimals", expected_decimals, f"call failed: {errors[dec_key]}")
        else:
            decimals = results.get(dec_key)
            _check(result, f"priceFeed({assets_str})/decimals", expected_decimals, decimals)

        ans_key = f"{feed_addr}/latestAnswer"
        if ans_key in errors:
            result.error(f"priceFeed({assets_str})/latestAnswer", "> 0", f"call failed: {errors[ans_key]}")
        else:
            answer = results.get(ans_key)
            if answer is not None and answer > 0:
                result.ok(f"priceFeed({assets_str})/latestAnswer", str(answer))
            else:
                result.error(f"priceFeed({assets_str})/latestAnswer", "> 0", str(answer))


def _check(
    result: VerificationResult, label: str, expected: Any, actual: Any
) -> None:
    # Config JSON stores large numeric values (WAD) as strings; web3.py
    # returns them as ints.  Coerce so the comparison works.
    if isinstance(expected, str):
        try:
            expected = int(expected)
        except ValueError:
            pass
    if expected == actual:
        result.ok(label, str(actual))
    else:
        result.error(label, expected, actual)


def load_method_selectors(sol_file: str, contract_name: str) -> dict[str, str]:
    """Return {function_name_prefix: "0x" + hex_selector} from Forge artifact.

    Keys are the function name (part before the first '('), values are the
    4-byte hex selector prefixed with "0x".  When multiple overloads share
    a name, all are returned keyed by their full signature.
    """
    path = ARTIFACTS_DIR / sol_file / f"{contract_name}.json"
    if not path.exists():
        sys.exit(f"Artifact not found: {path}\nRun `forge build` first.")
    with open(path) as f:
        mids = json.load(f)["methodIdentifiers"]
    result: dict[str, str] = {}
    for sig, sel_hex in mids.items():
        result[sig] = "0x" + sel_hex
    return result


def _selectors_for_names(
    method_ids: dict[str, str], names: list[str]
) -> list[tuple[str, bytes]]:
    """Given methodIdentifiers dict and a list of function name prefixes,
    return [(full_signature, 4-byte selector)] for each match."""
    out: list[tuple[str, bytes]] = []
    for name in names:
        matches = [
            (sig, bytes.fromhex(sel[2:]))
            for sig, sel in method_ids.items()
            if sig.split("(")[0] == name
        ]
        if not matches:
            print(f"  {RED}WARNING{RESET} selector not found for '{name}'")
        out.extend(matches)
    return out


# ---------------------------------------------------------------------------
# Role constants (mirroring Roles.sol)
# ---------------------------------------------------------------------------

ROLE_NAMES: dict[int, str] = {
    0: "ACCESS_MANAGER_DEFAULT_ADMIN",
    100: "HUB_DOMAIN_ADMIN_ROLE",
    101: "HUB_CONFIGURATOR_ROLE",
    102: "HUB_FEE_MINTER_ROLE",
    103: "HUB_DEFICIT_ELIMINATOR_ROLE",
    200: "HUB_CONFIGURATOR_DOMAIN_ADMIN_ROLE",
    300: "SPOKE_DOMAIN_ADMIN_ROLE",
    301: "SPOKE_CONFIGURATOR_ROLE",
    302: "SPOKE_USER_POSITION_UPDATER_ROLE",
    400: "SPOKE_CONFIGURATOR_DOMAIN_ADMIN_ROLE",
}

# role_id -> list of function name prefixes (matched against methodIdentifiers)
HUB_ROLE_FUNCTIONS: dict[int, list[str]] = {
    101: ["addAsset", "updateAssetConfig", "addSpoke", "updateSpokeConfig", "setInterestRateData"],
    102: ["mintFeeShares"],
    103: ["eliminateDeficit"],
}

HUB_CONFIGURATOR_ROLE_FUNCTIONS: dict[int, list[str]] = {
    200: [
        "addAsset", "addAssetWithDecimals", "updateLiquidityFee", "updateFeeReceiver",
        "updateFeeConfig", "updateInterestRateStrategy", "updateReinvestmentController",
        "resetAssetCaps", "deactivateAsset", "haltAsset", "addSpoke", "addSpokeToAssets",
        "updateSpokeActive", "updateSpokeHalted", "updateSpokeAddCap", "updateSpokeDrawCap",
        "updateSpokeRiskPremiumThreshold", "updateSpokeCaps", "deactivateSpoke", "haltSpoke",
        "resetSpokeCaps", "updateInterestRateData",
    ],
}

SPOKE_ROLE_FUNCTIONS: dict[int, list[str]] = {
    302: ["updateUserDynamicConfig", "updateUserRiskPremium"],
    301: [
        "updateLiquidationConfig", "addReserve", "updateReserveConfig",
        "updateDynamicReserveConfig", "addDynamicReserveConfig",
        "updatePositionManager", "updateReservePriceSource",
    ],
}

SPOKE_CONFIGURATOR_ROLE_FUNCTIONS: dict[int, list[str]] = {
    400: [
        "updateReservePriceSource", "updateLiquidationTargetHealthFactor",
        "updateHealthFactorForMaxBonus", "updateLiquidationBonusFactor",
        "updateLiquidationConfig", "addReserve", "updatePaused", "updateFrozen",
        "updateBorrowable", "updateReceiveSharesEnabled", "updateCollateralRisk",
        "addCollateralFactor", "updateCollateralFactor", "addMaxLiquidationBonus",
        "updateMaxLiquidationBonus", "addLiquidationFee", "updateLiquidationFee",
        "addDynamicReserveConfig", "updateDynamicReserveConfig", "pauseAllReserves",
        "freezeAllReserves", "pauseReserve", "freezeReserve", "updatePositionManager",
    ],
}


# ---------------------------------------------------------------------------
# TokenizationSpoke verification
# ---------------------------------------------------------------------------


def verify_tokenization_spokes(
    batch_mgr: BatchCallManager,
    caller: ContractCaller,
    cache: DeploymentCache,
    report: DeployReport,
    config: ConfigInput,
    result: VerificationResult,
    tokenization_report: dict | None = None,
) -> None:
    result.section("TokenizationSpoke Verification")

    if not config.tokens_by_key:
        print(f"  {YELLOW}SKIPPED{RESET} — no token addresses in config")
        return

    # Collect assets that have a tokenize section
    tokenize_entries: list[tuple[str, dict, str, int, str]] = []  # (label, tokenize_cfg, hub_addr, asset_id, hub_key)
    for entry in config.assets:
        tok = entry.get("tokenize")
        if not tok or not tok.get("name"):
            continue
        hub_key = entry["hubKey"]
        token_key = entry["tokenKey"]
        token_info = config.tokens_by_key.get(token_key)
        if not token_info or "address" not in token_info:
            continue
        hub_info = report.hub_by_label(hub_key)
        token_addr = Web3.to_checksum_address(token_info["address"])
        hub_addr = Web3.to_checksum_address(hub_info.address)
        asset_id = cache.asset_ids.get((hub_addr, token_addr))
        if asset_id is None:
            result.error(f"{hub_key}/{token_key}/tokenize", "valid assetId", "call failed")
            continue
        label = f"{hub_key}/{token_key}"
        tokenize_entries.append((label, tok, hub_addr, asset_id, hub_key))

    if not tokenize_entries:
        print("  No tokenization entries found in config.")
        return

    if tokenization_report is not None:
        _verify_tokenization_from_report(
            batch_mgr, caller, result, tokenize_entries, tokenization_report,
        )
    else:
        _verify_tokenization_by_discovery(
            batch_mgr, caller, result, tokenize_entries,
        )


def _verify_tokenization_from_report(
    batch_mgr: BatchCallManager,
    caller: ContractCaller,
    result: VerificationResult,
    tokenize_entries: list[tuple[str, dict, str, int, str]],
    tokenization_report: dict,
) -> None:
    """Verify TokenizationSpokes using known addresses from tokenization deploy report."""
    artifact = ARTIFACT_MAP["TokenizationSpoke"]

    # Resolve addresses and verify bytecode
    entry_addrs: list[tuple[str, dict, str, int, str]] = []  # (label, tok, hub_addr, asset_id, spoke_addr)
    for label, tok, hub_addr, asset_id, hub_key in tokenize_entries:
        token_key = label.split("/", 1)[1]
        hub_report = tokenization_report.get(hub_key)
        if hub_report is None:
            result.error(f"{label}/tokenize", "address in tokenization report", f"hub '{hub_key}' not found")
            continue
        spoke_addr_raw = hub_report.get(token_key)
        if spoke_addr_raw is None:
            result.error(f"{label}/tokenize", "address in tokenization report", f"token '{token_key}' not found")
            continue
        spoke_addr = Web3.to_checksum_address(spoke_addr_raw)

        # Bytecode: proxy
        _verify_single_bytecode(
            caller, result, f"{label}/tokenize/proxy",
            spoke_addr, artifact.sol_file, artifact.contract_name,
        )
        # Bytecode: implementation
        impl_addr = caller.get_implementation(spoke_addr)
        if impl_addr and int(impl_addr, 16) != 0:
            _verify_single_bytecode(
                caller, result, f"{label}/tokenize/impl",
                impl_addr, artifact.impl_sol_file, artifact.impl_contract_name,
            )
        else:
            result.error(f"{label}/tokenize/impl", "implementation address", "could not read proxy impl slot")

        entry_addrs.append((label, tok, hub_addr, asset_id, spoke_addr))

    if not entry_addrs:
        return

    # Batch name() and symbol() calls
    ns_calls: list[tuple[str, Any]] = []
    for label, _, _, _, spoke_addr in entry_addrs:
        tok_contract = batch_mgr.w3.eth.contract(
            address=spoke_addr, abi=caller.tokenization_spoke_abi,
        )
        ns_calls.append((f"{label}/name", tok_contract.functions.name()))
        ns_calls.append((f"{label}/symbol", tok_contract.functions.symbol()))

    ns_results, ns_errors = batch_mgr.execute(ns_calls)

    # Batch getSpokeConfig calls
    cfg_calls: list[tuple[str, Any]] = []
    for label, _, hub_addr, asset_id, spoke_addr in entry_addrs:
        hub_contract = batch_mgr.w3.eth.contract(
            address=hub_addr, abi=caller.hub_abi,
        )
        cfg_calls.append((
            f"{label}/spokeConfig",
            hub_contract.functions.getSpokeConfig(asset_id, spoke_addr),
        ))

    cfg_results, cfg_errors = batch_mgr.execute(cfg_calls)

    # Check results
    for label, tok, hub_addr, asset_id, spoke_addr in entry_addrs:
        expected_name = tok["name"]
        expected_symbol = tok["symbol"]
        expected_add_cap = tok["addCap"]

        # Name
        name_key = f"{label}/name"
        if name_key in ns_errors:
            result.error(f"{label}/tokenize/name", expected_name, f"call failed: {ns_errors[name_key]}")
        else:
            _check(result, f"{label}/tokenize/name", expected_name, ns_results.get(name_key))

        # Symbol
        symbol_key = f"{label}/symbol"
        if symbol_key in ns_errors:
            result.error(f"{label}/tokenize/symbol", expected_symbol, f"call failed: {ns_errors[symbol_key]}")
        else:
            _check(result, f"{label}/tokenize/symbol", expected_symbol, ns_results.get(symbol_key))

        # addCap from getSpokeConfig
        cfg_key = f"{label}/spokeConfig"
        if cfg_key in cfg_errors:
            result.error(f"{label}/tokenize/addCap", expected_add_cap, f"call failed: {cfg_errors[cfg_key]}")
        else:
            spoke_cfg = cfg_results.get(cfg_key)
            if spoke_cfg is not None:
                _check(result, f"{label}/tokenize/addCap", expected_add_cap, spoke_cfg[0])
            else:
                result.error(f"{label}/tokenize/addCap", expected_add_cap, "config call failed")


def _verify_tokenization_by_discovery(
    batch_mgr: BatchCallManager,
    caller: ContractCaller,
    result: VerificationResult,
    tokenize_entries: list[tuple[str, dict, str, int, str]],
) -> None:
    """Discover TokenizationSpokes by probing all spokes registered on Hub."""
    # Round 1: batch getSpokeCount for each (hub, assetId)
    count_calls: list[tuple[str, Any]] = []
    for label, _, hub_addr, asset_id, _ in tokenize_entries:
        hub_contract = batch_mgr.w3.eth.contract(
            address=hub_addr, abi=caller.hub_abi,
        )
        count_calls.append((label, hub_contract.functions.getSpokeCount(asset_id)))

    count_results, count_errors = batch_mgr.execute(count_calls)

    # Round 2: batch getSpokeAddress for each index
    addr_calls: list[tuple[str, Any]] = []
    addr_meta: list[tuple[str, dict, str, int, int]] = []  # label, tok, hub_addr, asset_id, index
    for label, tok, hub_addr, asset_id, _ in tokenize_entries:
        if label in count_errors:
            result.error(f"{label}/spokeCount", "spoke count", f"call failed: {count_errors[label]}")
            continue
        count = count_results.get(label)
        if count is None:
            result.error(f"{label}/spokeCount", "spoke count", "call failed")
            continue
        hub_contract = batch_mgr.w3.eth.contract(
            address=hub_addr, abi=caller.hub_abi,
        )
        for idx in range(count):
            key = f"{label}/spoke/{idx}"
            addr_calls.append((key, hub_contract.functions.getSpokeAddress(asset_id, idx)))
            addr_meta.append((label, tok, hub_addr, asset_id, idx))

    addr_results, addr_errors = batch_mgr.execute(addr_calls)

    # Round 3: Probe each spoke with sequential name()/symbol() calls.
    # Batching these is unsafe because name() reverts on non-TokenizationSpoke
    # contracts, which poisons the entire batch chunk.
    label_spokes: dict[str, list[str]] = {}
    spoke_names: dict[str, str] = {}   # spoke_addr -> name
    spoke_symbols: dict[str, str] = {} # spoke_addr -> symbol
    for label, tok, hub_addr, asset_id, idx in addr_meta:
        key = f"{label}/spoke/{idx}"
        if key in addr_errors:
            continue
        spoke_addr = addr_results.get(key)
        if spoke_addr is None:
            continue
        spoke_addr = Web3.to_checksum_address(spoke_addr)
        label_spokes.setdefault(label, []).append(spoke_addr)

        if spoke_addr not in spoke_names:
            name_val = caller.call_tokenization_spoke(spoke_addr, "name", silent=True)
            if name_val and isinstance(name_val, str):
                spoke_names[spoke_addr] = name_val
                symbol_val = caller.call_tokenization_spoke(spoke_addr, "symbol", silent=True)
                spoke_symbols[spoke_addr] = symbol_val if isinstance(symbol_val, str) else ""

    # Round 4: Batch getSpokeConfig only for confirmed TokenizationSpoke candidates
    config_calls: list[tuple[str, Any]] = []
    for label, tok, hub_addr, asset_id, _ in tokenize_entries:
        for spoke_addr in label_spokes.get(label, []):
            if spoke_addr in spoke_names:
                hub_contract = batch_mgr.w3.eth.contract(
                    address=hub_addr, abi=caller.hub_abi,
                )
                config_calls.append((
                    f"{label}/{spoke_addr}/config",
                    hub_contract.functions.getSpokeConfig(asset_id, spoke_addr),
                ))

    config_results, config_errors = batch_mgr.execute(config_calls)

    # Match discovered TokenizationSpokes against config expectations
    for label, tok, hub_addr, asset_id, _ in tokenize_entries:
        expected_name = tok["name"]
        expected_symbol = tok["symbol"]
        expected_add_cap = tok["addCap"]
        spokes = label_spokes.get(label, [])

        found = False
        for spoke_addr in spokes:
            name_val = spoke_names.get(spoke_addr)
            if not name_val:
                continue
            symbol_val = spoke_symbols.get(spoke_addr, "")

            if name_val == expected_name and symbol_val == expected_symbol:
                found = True
                result.ok(f"{label}/tokenize", f'name="{name_val}", symbol="{symbol_val}"')
                config_key = f"{label}/{spoke_addr}/config"
                if config_key in config_errors:
                    result.error(f"{label}/tokenize/addCap", expected_add_cap, "config call failed")
                else:
                    spoke_cfg = config_results.get(config_key)
                    if spoke_cfg is not None:
                        _check(result, f"{label}/tokenize/addCap", expected_add_cap, spoke_cfg[0])
                    else:
                        result.error(f"{label}/tokenize/addCap", expected_add_cap, "config call failed")
                break

        if not found:
            result.error(
                f"{label}/tokenize",
                f'TokenizationSpoke name="{expected_name}" symbol="{expected_symbol}"',
                f"not found among {len(spokes)} spokes",
            )


# ---------------------------------------------------------------------------
# Role verification
# ---------------------------------------------------------------------------


def _build_selector_role_map(
    method_ids: dict[str, str],
    role_functions: dict[int, list[str]],
) -> list[tuple[str, bytes, int]]:
    """Return [(function_signature, 4-byte selector, expected_role_id)]."""
    out: list[tuple[str, bytes, int]] = []
    for role_id, fn_names in role_functions.items():
        for sig, sel_bytes in _selectors_for_names(method_ids, fn_names):
            out.append((sig, sel_bytes, role_id))
    return out


def verify_roles(
    batch_mgr: BatchCallManager,
    caller: ContractCaller,
    report: DeployReport,
    config: ConfigInput,
    result: VerificationResult,
) -> None:
    result.section("Role Configuration")

    am_addr = Web3.to_checksum_address(report.access_manager)
    am_contract = batch_mgr.w3.eth.contract(
        address=am_addr, abi=caller.access_manager_abi,
    )

    hub_mids = load_method_selectors("IHub.sol", "IHub")
    hub_cfg_mids = load_method_selectors("IHubConfigurator.sol", "IHubConfigurator")
    spoke_mids = load_method_selectors("ISpoke.sol", "ISpoke")
    spoke_cfg_mids = load_method_selectors("ISpokeConfigurator.sol", "ISpokeConfigurator")

    hub_sel_roles = _build_selector_role_map(hub_mids, HUB_ROLE_FUNCTIONS)
    hub_cfg_sel_roles = _build_selector_role_map(hub_cfg_mids, HUB_CONFIGURATOR_ROLE_FUNCTIONS)
    spoke_sel_roles = _build_selector_role_map(spoke_mids, SPOKE_ROLE_FUNCTIONS)
    spoke_cfg_sel_roles = _build_selector_role_map(spoke_cfg_mids, SPOKE_CONFIGURATOR_ROLE_FUNCTIONS)

    # --- Part A: Selector-to-Role Mapping ---
    print(f"\n{BOLD}--- Selector-to-Role Mapping ---{RESET}\n")

    role_calls: list[tuple[str, Any]] = []
    role_expectations: list[tuple[str, int]] = []  # (key, expected_role)

    # Hub targets
    for hub_info in report.hubs:
        target = Web3.to_checksum_address(hub_info.address)
        for sig, sel_bytes, expected_role in hub_sel_roles:
            fn_name = sig.split("(")[0]
            key = f"Hub({hub_info.label})/{fn_name}"
            role_calls.append((
                key,
                am_contract.functions.getTargetFunctionRole(target, sel_bytes),
            ))
            role_expectations.append((key, expected_role))

    # HubConfigurator target
    hc_target = Web3.to_checksum_address(report.hub_configurator)
    for sig, sel_bytes, expected_role in hub_cfg_sel_roles:
        fn_name = sig.split("(")[0]
        key = f"HubConfigurator/{fn_name}"
        role_calls.append((
            key,
            am_contract.functions.getTargetFunctionRole(hc_target, sel_bytes),
        ))
        role_expectations.append((key, expected_role))

    # Spoke targets
    for spoke_info in report.spokes:
        target = Web3.to_checksum_address(spoke_info.proxy)
        for sig, sel_bytes, expected_role in spoke_sel_roles:
            fn_name = sig.split("(")[0]
            key = f"Spoke({spoke_info.label})/{fn_name}"
            role_calls.append((
                key,
                am_contract.functions.getTargetFunctionRole(target, sel_bytes),
            ))
            role_expectations.append((key, expected_role))

    # SpokeConfigurator target
    sc_target = Web3.to_checksum_address(report.spoke_configurator)
    for sig, sel_bytes, expected_role in spoke_cfg_sel_roles:
        fn_name = sig.split("(")[0]
        key = f"SpokeConfigurator/{fn_name}"
        role_calls.append((
            key,
            am_contract.functions.getTargetFunctionRole(sc_target, sel_bytes),
        ))
        role_expectations.append((key, expected_role))

    role_results, role_errors = batch_mgr.execute(role_calls)

    for key, expected_role in role_expectations:
        if key in role_errors:
            result.error(key, f"role {expected_role}", f"call failed: {role_errors[key]}")
        else:
            actual = role_results.get(key)
            role_name = ROLE_NAMES.get(expected_role, str(expected_role))
            if actual == expected_role:
                result.ok(key, f"{role_name} ({expected_role})")
            else:
                actual_name = ROLE_NAMES.get(actual, str(actual)) if actual is not None else "None"
                result.error(
                    key,
                    f"{role_name} ({expected_role})",
                    f"{actual_name} ({actual})",
                )

    # --- Part B: Structural Role Grants ---
    print(f"\n{BOLD}--- Structural Role Grants ---{RESET}\n")

    grant_calls: list[tuple[str, Any]] = []
    grant_expectations: list[tuple[str, str]] = []

    # HUB_CONFIGURATOR_ROLE (101) -> HubConfigurator should be member
    grant_calls.append((
        "HUB_CONFIGURATOR_ROLE/HubConfigurator",
        am_contract.functions.hasRole(101, hc_target),
    ))
    grant_expectations.append(("HUB_CONFIGURATOR_ROLE/HubConfigurator", "HubConfigurator is member"))

    # SPOKE_CONFIGURATOR_ROLE (301) -> SpokeConfigurator should be member
    grant_calls.append((
        "SPOKE_CONFIGURATOR_ROLE/SpokeConfigurator",
        am_contract.functions.hasRole(301, sc_target),
    ))
    grant_expectations.append(("SPOKE_CONFIGURATOR_ROLE/SpokeConfigurator", "SpokeConfigurator is member"))

    grant_results, grant_errors = batch_mgr.execute(grant_calls)

    for key, description in grant_expectations:
        if key in grant_errors:
            result.error(key, description, f"call failed: {grant_errors[key]}")
        else:
            val = grant_results.get(key)
            # hasRole returns (isMember, executionDelay)
            is_member = val[0] if isinstance(val, (tuple, list)) else val
            if is_member:
                result.ok(key, description)
            else:
                result.error(key, description, "not a member")

    # --- Part C: Print All Role Members (informational) ---
    print(f"\n{BOLD}--- Role Members (informational) ---{RESET}\n")

    # Build address -> label reverse lookup from the deploy report
    addr_labels: dict[str, str] = {}
    for label, addr in report.all_report_keys():
        addr_labels[Web3.to_checksum_address(addr)] = label

    # Get total role count
    role_count = caller.call_access_manager(am_addr, "getRoleCount")
    if role_count is None or role_count == 0:
        print("  Could not retrieve role count.")
        return

    # Get all role IDs
    roles = caller.call_access_manager(am_addr, "getRoles", 0, role_count)
    if roles is None:
        print("  Could not retrieve roles.")
        return

    # Batch: get member count for each role
    member_count_calls: list[tuple[str, Any]] = []
    for role_id in roles:
        key = f"role_{role_id}/memberCount"
        member_count_calls.append((
            key,
            am_contract.functions.getRoleMemberCount(role_id),
        ))

    mc_results, mc_errors = batch_mgr.execute(member_count_calls)

    # Batch: get members for each role
    member_calls: list[tuple[str, Any]] = []
    role_member_counts: dict[int, int] = {}
    for role_id in roles:
        key = f"role_{role_id}/memberCount"
        if key in mc_errors:
            continue
        count = mc_results.get(key, 0)
        if count and count > 0:
            role_member_counts[role_id] = count
            member_calls.append((
                f"role_{role_id}/members",
                am_contract.functions.getRoleMembers(role_id, 0, count),
            ))

    m_results, m_errors = batch_mgr.execute(member_calls)

    for role_id in roles:
        role_name = ROLE_NAMES.get(role_id, f"UNKNOWN_ROLE")
        print(f"  Role {role_id} ({role_name}):")
        mkey = f"role_{role_id}/members"
        if mkey in m_errors:
            print(f"    (error fetching members: {m_errors[mkey]})")
        elif mkey in m_results:
            members = m_results[mkey]
            if members:
                for addr in members:
                    label = addr_labels.get(Web3.to_checksum_address(addr), "")
                    suffix = f" ({label})" if label else ""
                    print(f"    - {addr}{suffix}")
            else:
                print(f"    (no members)")
        else:
            print(f"    (no members)")

    # --- Part D: Role Labels ---
    print(f"\n{BOLD}--- Role Labels ---{RESET}\n")

    # Role 0 (ADMIN_ROLE) cannot be labeled by the AccessManager contract
    labelable_roles = {rid: name for rid, name in ROLE_NAMES.items() if rid != 0}

    # Batch 1: check which roles are labeled
    labeled_calls: list[tuple[str, Any]] = []
    for role_id in labelable_roles:
        key = f"role_{role_id}/isLabeled"
        labeled_calls.append((
            key,
            am_contract.functions.isRoleLabeled(role_id),
        ))

    labeled_results, labeled_errors = batch_mgr.execute(labeled_calls)

    # Batch 2: get labels for roles that are labeled
    label_calls: list[tuple[str, Any]] = []
    for role_id in labelable_roles:
        key = f"role_{role_id}/isLabeled"
        if key in labeled_errors:
            continue
        if labeled_results.get(key) is True:
            label_key = f"role_{role_id}/label"
            label_calls.append((
                label_key,
                am_contract.functions.getLabelOfRole(role_id),
            ))

    label_results, label_errors = batch_mgr.execute(label_calls)

    # Verify
    for role_id, expected_label in labelable_roles.items():
        is_labeled_key = f"role_{role_id}/isLabeled"
        if is_labeled_key in labeled_errors:
            result.error(
                f"role {role_id}/label",
                expected_label,
                f"call failed: {labeled_errors[is_labeled_key]}",
            )
            continue

        if labeled_results.get(is_labeled_key) is not True:
            result.error(f"role {role_id}/label", expected_label, "role not labeled")
            continue

        label_key = f"role_{role_id}/label"
        if label_key in label_errors:
            result.error(
                f"role {role_id}/label",
                expected_label,
                f"call failed: {label_errors[label_key]}",
            )
        else:
            actual_label = label_results.get(label_key)
            _check(result, f"role {role_id}/label", expected_label, actual_label)


def verify_position_managers_and_gateways(
    batch_mgr: BatchCallManager,
    caller: ContractCaller,
    report: DeployReport,
    result: VerificationResult,
) -> None:
    result.section("Position Manager & Gateway Verification")

    pm_gw_entries: list[tuple[str, str]] = []
    for name, addr in [
        ("GiverPositionManager", report.giver_position_manager),
        ("TakerPositionManager", report.taker_position_manager),
        ("ConfigPositionManager", report.config_position_manager),
        ("SignatureGateway", report.signature_gateway),
        ("NativeTokenGateway", report.native_token_gateway),
    ]:
        if addr:
            pm_gw_entries.append((name, Web3.to_checksum_address(addr)))

    if not pm_gw_entries:
        print("  No position managers or gateways found in report.")
        return

    if not report.spokes:
        print("  No spokes found in report.")
        return

    # --- Spoke-side: isPositionManagerActive ---
    print(f"\n{BOLD}--- Spoke-side (isPositionManagerActive) ---{RESET}\n")

    active_calls: list[tuple[str, Any]] = []
    for spoke_info in report.spokes:
        spoke_addr = Web3.to_checksum_address(spoke_info.proxy)
        spoke_contract = batch_mgr.w3.eth.contract(
            address=spoke_addr, abi=caller.spoke_abi,
        )
        for pm_name, pm_addr in pm_gw_entries:
            key = f"{spoke_info.label}/{pm_name}"
            active_calls.append((
                key,
                spoke_contract.functions.isPositionManagerActive(pm_addr),
            ))

    active_results, active_errors = batch_mgr.execute(active_calls)

    for spoke_info in report.spokes:
        for pm_name, _ in pm_gw_entries:
            key = f"{spoke_info.label}/{pm_name}"
            if key in active_errors:
                result.error(key, "active", f"call failed: {active_errors[key]}")
            else:
                val = active_results.get(key)
                if val is True:
                    result.ok(key, "active")
                else:
                    result.error(key, "active", f"{val}")

    # --- PM/Gateway-side: isSpokeRegistered ---
    print(f"\n{BOLD}--- PM/Gateway-side (isSpokeRegistered) ---{RESET}\n")

    reg_calls: list[tuple[str, Any]] = []
    for pm_name, pm_addr in pm_gw_entries:
        pm_contract = batch_mgr.w3.eth.contract(
            address=pm_addr, abi=caller.position_manager_base_abi,
        )
        for spoke_info in report.spokes:
            spoke_addr = Web3.to_checksum_address(spoke_info.proxy)
            key = f"{pm_name}/{spoke_info.label}"
            reg_calls.append((
                key,
                pm_contract.functions.isSpokeRegistered(spoke_addr),
            ))

    reg_results, reg_errors = batch_mgr.execute(reg_calls)

    for pm_name, _ in pm_gw_entries:
        for spoke_info in report.spokes:
            key = f"{pm_name}/{spoke_info.label}"
            if key in reg_errors:
                result.error(key, "registered", f"call failed: {reg_errors[key]}")
            else:
                val = reg_results.get(key)
                if val is True:
                    result.ok(key, "registered")
                else:
                    result.error(key, "registered", f"{val}")


def verify_etherscan_source(
    caller: ContractCaller,
    report: DeployReport,
    result: VerificationResult,
    api_key: str,
    chain_id: int = 1,
) -> dict[str, dict]:
    """Check that every deployed address is source-verified on Etherscan.

    For proxy contracts the implementation is also checked.  Returns a cache
    mapping ``address -> etherscan result dict`` for reuse by compiler checks.
    """
    result.section("Etherscan Source Verification")

    cache: dict[str, dict] = {}

    def _check_addr(label: str, addr: str) -> None:
        addr = Web3.to_checksum_address(addr)
        if addr in cache:
            src = cache[addr]
        else:
            src = _etherscan_get_source(addr, api_key, chain_id)
            if src is not None:
                cache[addr] = src
        if src is None:
            result.error(label, "Etherscan response", f"no data for {addr}")
        elif not src.get("SourceCode"):
            result.error(label, "source verified", f"not verified at {addr}")
        else:
            result.ok(label, addr)

    for name, addr in report.all_addresses():
        _check_addr(name, addr)

        suffix = name.rsplit("/", 1)[-1]
        artifact = ARTIFACT_MAP.get(suffix)
        if artifact and artifact.impl_sol_file:
            impl_addr = caller.get_implementation(addr)
            _check_addr(f"{name}/Implementation", impl_addr)

    return cache


def verify_etherscan_compiler(
    caller: ContractCaller,
    report: DeployReport,
    result: VerificationResult,
    etherscan_cache: dict[str, dict],
) -> None:
    """Compare Etherscan-reported compiler settings against expected values."""
    result.section("Compiler Settings Verification (Etherscan)")

    checked: set[str] = set()

    def _check_compiler(label: str, addr: str, contract_name: str) -> None:
        addr = Web3.to_checksum_address(addr)
        if addr in checked:
            return
        checked.add(addr)

        src = etherscan_cache.get(addr)
        if src is None or not src.get("SourceCode"):
            result.error(label, "compiler data", f"no verified source for {addr}")
            return

        expected = _expected_compiler(contract_name)

        compiler_ver = src.get("CompilerVersion", "")
        if f"v{expected.solc}" in compiler_ver:
            result.ok(f"{label}/solc", compiler_ver)
        else:
            result.error(f"{label}/solc", f"v{expected.solc}", compiler_ver)

        opt_used = src.get("OptimizationUsed", "0")
        if opt_used == "1":
            result.ok(f"{label}/optimizer", "enabled")
        else:
            result.error(f"{label}/optimizer", "1 (enabled)", opt_used)

        runs = int(src.get("Runs", "0"))
        # Etherscan may truncate very large optimizer_runs to 32-bit
        expected_runs = expected.optimizer_runs
        expected_runs_truncated = expected_runs & 0xFFFFFFFF
        if runs == expected_runs or runs == expected_runs_truncated:
            result.ok(f"{label}/runs", str(runs))
        else:
            result.error(f"{label}/runs", str(expected_runs), str(runs))

        evm = src.get("EVMVersion", "").lower()
        if evm == "default":
            # Etherscan reports "default" when the EVM version matches the
            # compiler default; for solc 0.8.28 that is "cancun".
            evm = "cancun"
        if evm == expected.evm_version:
            result.ok(f"{label}/evmVersion", evm)
        else:
            result.error(f"{label}/evmVersion", expected.evm_version, evm)

        # Extract viaIR from Standard JSON input when available
        actual_via_ir = False
        source_code = src.get("SourceCode", "")
        if source_code.startswith("{{"):
            try:
                inner = json.loads(source_code[1:-1])
                actual_via_ir = inner.get("settings", {}).get("viaIR", False)
            except (json.JSONDecodeError, KeyError):
                pass
        if actual_via_ir == expected.via_ir:
            result.ok(f"{label}/viaIR", str(actual_via_ir))
        else:
            result.error(f"{label}/viaIR", str(expected.via_ir), str(actual_via_ir))

    for name, addr in report.all_addresses():
        suffix = name.rsplit("/", 1)[-1]
        artifact = ARTIFACT_MAP.get(suffix)
        if artifact is None:
            continue

        _check_compiler(name, addr, artifact.contract_name)

        if artifact.impl_sol_file:
            impl_addr = caller.get_implementation(addr)
            _check_compiler(
                f"{name}/Implementation", impl_addr, artifact.impl_contract_name,
            )


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Verify an Aave V4 deployment against its configuration."
    )
    parser.add_argument("--rpc-url", required=True, help="RPC endpoint URL")
    parser.add_argument(
        "--report", required=True, help="Path to deployment report JSON"
    )
    parser.add_argument("--config", required=True, help="Path to config input (.xlsx or .json)")
    parser.add_argument(
        "--tokenization-report", default=None,
        help="Path to tokenization deployment report JSON",
    )
    parser.add_argument(
        "--tokens", default=None,
        help="Path to JSON file with token addresses and price feeds",
    )
    parser.add_argument(
        "--etherscan-api-key", default=None,
        help="Etherscan API key for source & compiler verification (mainnet only)",
    )
    args = parser.parse_args()

    with open(args.report) as f:
        report = DeployReport.from_json(json.load(f))

    if args.config.endswith(".xlsx"):
        config = ConfigInput(load_config_from_xlsx(args.config))
    else:
        with open(args.config) as f:
            config = ConfigInput(json.load(f))

    if args.tokens:
        with open(args.tokens) as f:
            config.tokens_by_key = json.load(f)

    tokenization_report = None
    if args.tokenization_report:
        with open(args.tokenization_report) as f:
            tokenization_report = json.load(f)

    w3 = Web3(Web3.HTTPProvider(args.rpc_url))
    if not w3.is_connected():
        print(f"{RED}ERROR{RESET}: Cannot connect to {args.rpc_url}")
        sys.exit(1)

    print(f"Connected to chain {w3.eth.chain_id}")

    caller = ContractCaller(w3)
    result = VerificationResult()
    batch_mgr = BatchCallManager(w3)

    # Bytecode verification stays sequential (uses get_code/get_storage_at)
    verify_bytecode(caller, report, result)

    chain_id = w3.eth.chain_id
    if chain_id == 1:
        if args.etherscan_api_key:
            etherscan_cache = verify_etherscan_source(
                caller, report, result, args.etherscan_api_key, chain_id,
            )
            verify_etherscan_compiler(caller, report, result, etherscan_cache)
        else:
            print(
                f"\n{BOLD}=== Etherscan Verification ==={RESET}\n"
                f"  {RED}WARNING{RESET} --etherscan-api-key not provided; "
                "skipping source & compiler verification."
            )

    # Pre-fetch shared lookups (asset_ids, reserve_ids) in 2 batch round-trips
    cache = prefetch_shared_data(batch_mgr, caller, report, config)

    verify_hub_assets(batch_mgr, caller, cache, report, config, result)
    verify_spoke_registrations(batch_mgr, caller, cache, report, config, result)
    verify_reserves(batch_mgr, caller, cache, report, config, result)
    verify_liquidation_configs(batch_mgr, caller, report, config, result)
    verify_oracles(batch_mgr, caller, cache, report, config, result)
    verify_oracle_wiring(batch_mgr, caller, report, result)
    verify_price_feeds(batch_mgr, caller, report, config, result)
    verify_tokenization_spokes(batch_mgr, caller, cache, report, config, result, tokenization_report)
    verify_roles(batch_mgr, caller, report, config, result)
    # verify_position_managers_and_gateways(batch_mgr, caller, report, result)

    sys.exit(result.summary())


if __name__ == "__main__":
    main()
